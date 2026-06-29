"""Inject Claude Prompt 2 commentary into SMPL_MDA_Package Excel template."""

from __future__ import annotations

import calendar
import io
import os
import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell

from app.core.config import _BACKEND_ROOT
from app.services.reporting.export.mda_excel_values import (
    FMT_HEADCOUNT,
    FMT_MONEY_M,
    FMT_PCT,
    FMT_VAR_M,
    parse_deck_money_to_m,
    parse_deck_pct_to_ratio,
)
from app.services.reporting.export.schemas import ReportingBundle

VARIANCE_SHEET = "Variance Commentary"
RISKS_SHEET = "Risks & Opportunities"

VARIANCE_ROW_MAP: dict[str, int] = {
    "vc_arr_ending": 3,
    "vc_net_new_arr": 4,
    "vc_revenue": 5,
    "vc_gross_margin": 6,
    "vc_ebitda": 7,
    "vc_sm_pipeline": 8,
    "vc_cash": 9,
    "vc_headcount": 10,
}

RISKS_ROW_MAP: dict[str, int] = {
    "ro_gtm_paid_search": 3,
    "ro_retention_smb": 4,
    "ro_pipeline": 5,
    "ro_cash_h2": 6,
    "ro_opp_gtm": 7,
    "ro_opp_expansion": 8,
    "ro_opp_margin": 9,
    "ro_opp_cash": 10,
}

COMMENTARY_HEADERS = (
    "Period vs Budget Commentary",
    "QTD vs Budget Commentary",
    "YTD vs Budget Commentary",
)

HEADER_ROW = 2
DATA_STYLE_ROW = 3
COMMENTARY_STYLE_ROW = 3

# Commentary columns sit after the FY Var % block (col 29) on monthly grid sheets.
MONTHLY_GRID_HEADER_ROW = 3
MONTHLY_COMMENTARY_START_COL = 30

SHEET_COMMENTARY_LAYOUT: dict[str, dict[str, int]] = {
    "ARR Waterfall": {"header_row": MONTHLY_GRID_HEADER_ROW, "start_col": MONTHLY_COMMENTARY_START_COL},
    "Income Statement": {"header_row": MONTHLY_GRID_HEADER_ROW, "start_col": MONTHLY_COMMENTARY_START_COL},
    "Cash Forecast": {"header_row": MONTHLY_GRID_HEADER_ROW, "start_col": MONTHLY_COMMENTARY_START_COL},
    "Cash Flow Statement": {"header_row": MONTHLY_GRID_HEADER_ROW, "start_col": MONTHLY_COMMENTARY_START_COL},
    "Q2 vs Budget": {"header_row": 4, "start_col": 10},
    "GTM Review": {"header_row": 2, "start_col": 7},
}

GTM_DATA_COL_WIDTHS = {"A": 22.0, "B": 12.0, "C": 14.0, "D": 12.0, "E": 12.0, "F": 10.0}
GTM_COMMENTARY_COL_WIDTH = 28.0


def _resolve_write_cell(ws, row: int, col: int) -> tuple[int, int]:
    """Merged cells are read-only except the top-left anchor."""
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return row, col
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col
    return row, col


def _set_cell(ws, row: int, col: int, value: Any) -> None:
    row, col = _resolve_write_cell(ws, row, col)
    ws.cell(row, col, value)


@dataclass(frozen=True)
class VarianceSheetLayout:
    """Dynamic column positions on the Variance Commentary tab."""

    metric_start_col: int
    metric_end_col: int
    commentary_start_col: int

    @property
    def commentary_cols(self) -> tuple[int, int, int]:
        c = self.commentary_start_col
        return (c, c + 1, c + 2)


def _detect_variance_layout(ws) -> VarianceSheetLayout:
    """Find metric vs commentary columns from the header row (adapts as periods change)."""
    metric_end = 2
    commentary_start: int | None = None
    for col in range(1, ws.max_column + 5):
        raw = ws.cell(HEADER_ROW, col).value
        if raw is None:
            if col > metric_end + 12:
                break
            continue
        label = str(raw).strip()
        lower = label.lower()
        if "commentary" in lower:
            commentary_start = col
            break
        metric_end = col
    if commentary_start is None:
        commentary_start = metric_end + 1
    return VarianceSheetLayout(
        metric_start_col=3,
        metric_end_col=max(metric_end, 11),
        commentary_start_col=commentary_start,
    )


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def _style_template_cell(ws, row: int, col: int):
    """Prefer an existing body cell; fall back to header row for new columns."""
    if ws.cell(row, col).value is not None or row == HEADER_ROW:
        return ws.cell(row, col)
    return ws.cell(DATA_STYLE_ROW, col)


def _extend_title_merge(ws, end_col: int) -> None:
    from app.services.reporting.export.mda_excel_values import extend_title_merge

    extend_title_merge(ws, end_col)


def _metric_value_and_format(value: Any, *, value_kind: str) -> tuple[Any, str]:
    if value_kind == "headcount":
        raw = str(value or "").strip()
        if raw in ("", "—", "n/a"):
            return None, FMT_HEADCOUNT
        try:
            return int(raw.replace(",", "")), FMT_HEADCOUNT
        except ValueError:
            return raw, FMT_HEADCOUNT
    if value_kind == "percent":
        return parse_deck_pct_to_ratio(value), FMT_PCT
    if isinstance(value, (int, float)):
        return float(value), FMT_MONEY_M
    text = str(value or "").strip()
    if not text or text in ("—", "n/a"):
        return None, FMT_MONEY_M
    if text.startswith("+") or (text.startswith("$") and "K" in text.upper()):
        return parse_deck_money_to_m(text), FMT_VAR_M
    return parse_deck_money_to_m(text), FMT_MONEY_M


def _apply_metric_cell(ws, row: int, col: int, value: Any, *, value_kind: str) -> None:
    row, col = _resolve_write_cell(ws, row, col)
    numeric, num_fmt = _metric_value_and_format(value, value_kind=value_kind)
    style_src = _style_template_cell(ws, DATA_STYLE_ROW, col)
    cell = ws.cell(row, col, numeric)
    _copy_cell_style(style_src, cell)
    cell.number_format = num_fmt


def _apply_commentary_cell(ws, row: int, col: int, text: str) -> None:
    row, col = _resolve_write_cell(ws, row, col)
    style_src = ws.cell(COMMENTARY_STYLE_ROW, col)
    if style_src.value is None and col > 12:
        style_src = ws.cell(COMMENTARY_STYLE_ROW, 12)
    cell = ws.cell(row, col, text)
    _copy_cell_style(style_src, cell)
    from openpyxl.styles import Alignment

    cell.alignment = Alignment(
        horizontal=style_src.alignment.horizontal or "center",
        vertical="top",
        wrap_text=True,
    )
    cell.number_format = "@"

CHAR_LIMITS: dict[str, dict[str, int]] = {
    "variance_commentary": {"period_vs_budget": 400, "qtd_vs_budget": 400, "ytd_vs_budget": 400},
    "q2_vs_budget": {"period_vs_budget": 300, "qtd_vs_budget": 300, "ytd_vs_budget": 300},
    "arr_waterfall": {"period_vs_budget": 300, "qtd_vs_budget": 300, "ytd_vs_budget": 300},
    "income_statement": {"period_vs_budget": 200, "qtd_vs_budget": 200, "ytd_vs_budget": 200},
    "cash_forecast": {"period_vs_budget": 200, "qtd_vs_budget": 200, "ytd_vs_budget": 200},
    "cash_flow_statement": {"period_vs_budget": 200, "qtd_vs_budget": 200, "ytd_vs_budget": 200},
    "gtm_review": {"period_vs_budget": 175, "qtd_vs_budget": 175, "ytd_vs_budget": 175},
    "headcount": {"period_vs_budget": 200, "qtd_vs_budget": 200, "ytd_vs_budget": 200},
    "risks_and_opportunities": {"description": 350, "recommended_action": 200},
}


def resolve_mda_package_template() -> Path | None:
    """Locate SMPL_MDA_Package reference workbook."""
    candidates: list[Path] = []
    env = os.environ.get("MDA_PACKAGE_TEMPLATE", "").strip()
    if env:
        candidates.append(Path(env))
    repo = _BACKEND_ROOT.parent
    candidates.extend(
        [
            _BACKEND_ROOT / "templates" / "mda" / "SMPL_MDA_Package_Template.xlsx",
            repo / "frontend" / "public" / "board" / "exports" / "SMPL_MDA_Package_June2026.xlsx",
            Path.home() / "Downloads" / "SMPL_MDA_Package_June2026.xlsx",
            Path.home() / "OneDrive" / "SMPL_MDA_Package_June2026.xlsx",
        ]
    )
    for path in candidates:
        if path.is_file():
            return path
    return None


def _truncate_at_sentence(text: str, limit: int) -> str:
    text = (text or "").strip().replace("\n", " ")
    if len(text) <= limit:
        return text
    clipped = text[: limit - 1].rstrip()
    for sep in (". ", "; "):
        idx = clipped.rfind(sep)
        if idx > limit * 0.55:
            return clipped[: idx + 1].strip()
    if " " in clipped:
        clipped = clipped.rsplit(" ", 1)[0]
    return clipped.rstrip(".,;:") + "…"


def trim_commentary_response(response: dict[str, Any]) -> dict[str, Any]:
    """Enforce per-sheet character limits on Claude JSON."""
    out: dict[str, Any] = {}
    for sheet, rows in response.items():
        if not isinstance(rows, dict):
            continue
        limits = CHAR_LIMITS.get(sheet, {})
        sheet_out: dict[str, Any] = {}
        for row_id, cols in rows.items():
            if not isinstance(cols, dict):
                continue
            if sheet == "risks_and_opportunities":
                sheet_out[row_id] = {
                    k: _truncate_at_sentence(str(cols.get(k, "")), limits.get(k, 350))
                    for k in ("description", "recommended_action")
                }
            else:
                sheet_out[row_id] = {
                    k: _truncate_at_sentence(str(cols.get(k, "")), limits.get(k, 300))
                    for k in ("period_vs_budget", "qtd_vs_budget", "ytd_vs_budget")
                }
        out[sheet] = sheet_out
    return out


def _month_label(period: str) -> str:
    year, month = period.split("-")
    return calendar.month_name[int(month)]


def _quarter_label(period: str) -> str:
    month = int(period.split("-")[1])
    return f"Q{(month - 1) // 3 + 1}"


def apply_variance_metrics(ws, display: dict[str, Any], close_period: str) -> VarianceSheetLayout:
    """Refresh Variance Commentary metric columns from warehouse payload."""
    layout = _detect_variance_layout(ws)
    month = _month_label(close_period)
    q = _quarter_label(close_period)

    metric_headers = [
        f"{month} Actual",
        f"{month} Budget",
        f"{month} Var",
        f"{q} Actual",
        f"{q} Budget",
        f"{q} Var",
        "YTD Actual",
        "YTD Budget",
        "YTD Var",
    ]
    _set_cell(ws, 1, 1, display.get("sheet_title", ws.cell(1, 1).value))
    _set_cell(ws, HEADER_ROW, 1, "Category")
    _set_cell(ws, HEADER_ROW, 2, "Metric")
    header_style = ws.cell(HEADER_ROW, 1)
    for offset, header in enumerate(metric_headers):
        col = layout.metric_start_col + offset
        row, col = _resolve_write_cell(ws, HEADER_ROW, col)
        cell = ws.cell(row, col, header)
        _copy_cell_style(header_style, cell)
    commentary_style = ws.cell(HEADER_ROW, layout.commentary_start_col)
    if commentary_style.value is None:
        commentary_style = ws.cell(HEADER_ROW, 12)
    for offset, header in enumerate(COMMENTARY_HEADERS):
        col = layout.commentary_start_col + offset
        row, col = _resolve_write_cell(ws, HEADER_ROW, col)
        cell = ws.cell(row, col, header)
        _copy_cell_style(commentary_style, cell)

    for row_def in display.get("rows") or []:
        row_id = row_def.get("row_id")
        excel_row = VARIANCE_ROW_MAP.get(str(row_id))
        if not excel_row:
            continue
        value_kind = str(row_def.get("value_kind") or "money")
        ws.cell(excel_row, 1, row_def.get("category", ""))
        ws.cell(excel_row, 2, row_def.get("metric", ""))
        cm = row_def.get("cm") or {}
        qtd = row_def.get("qtd") or {}
        ytd = row_def.get("ytd") or {}
        values = [
            cm.get("actual", ""),
            cm.get("budget", ""),
            cm.get("var", ""),
            qtd.get("actual", ""),
            qtd.get("budget", ""),
            qtd.get("var", ""),
            ytd.get("actual", ""),
            ytd.get("budget", ""),
            ytd.get("var", ""),
        ]
        for offset, val in enumerate(values):
            col = layout.metric_start_col + offset
            _apply_metric_cell(ws, excel_row, col, val, value_kind=value_kind)

    from openpyxl.utils import get_column_letter

    title_end = layout.commentary_start_col + 2
    _extend_title_merge(ws, title_end)
    for col in layout.commentary_cols:
        ws.column_dimensions[get_column_letter(col)].width = GTM_COMMENTARY_COL_WIDTH

    return layout


def inject_commentary(
    workbook: Any,
    commentary: dict[str, Any],
    *,
    close_period: str,
    display: dict[str, Any] | None = None,
) -> None:
    """Write commentary strings into template workbook."""
    if VARIANCE_SHEET in workbook.sheetnames:
        ws = workbook[VARIANCE_SHEET]
        layout = None
        if display:
            layout = apply_variance_metrics(ws, display, close_period)
        else:
            layout = _detect_variance_layout(ws)
        vc = commentary.get("variance_commentary") or {}
        c1, c2, c3 = layout.commentary_cols
        for row_id, excel_row in VARIANCE_ROW_MAP.items():
            cols = vc.get(row_id) or {}
            _apply_commentary_cell(ws, excel_row, c1, cols.get("period_vs_budget", ""))
            _apply_commentary_cell(ws, excel_row, c2, cols.get("qtd_vs_budget", ""))
            _apply_commentary_cell(ws, excel_row, c3, cols.get("ytd_vs_budget", ""))

    if RISKS_SHEET in workbook.sheetnames:
        ws = workbook[RISKS_SHEET]
        ro = commentary.get("risks_and_opportunities") or {}
        for row_id, excel_row in RISKS_ROW_MAP.items():
            cols = ro.get(row_id) or {}
            _set_cell(ws, excel_row, 4, cols.get("description", ""))
            _set_cell(ws, excel_row, 6, cols.get("recommended_action", ""))

    _inject_sheet_commentary_cols(
        workbook, "ARR Waterfall", commentary.get("arr_waterfall") or {}, ws_name="ARR Waterfall"
    )
    _inject_sheet_commentary_cols(
        workbook, "Income Statement", commentary.get("income_statement") or {}, ws_name="Income Statement"
    )
    _inject_sheet_commentary_cols(
        workbook, "Cash Forecast", commentary.get("cash_forecast") or {}, ws_name="Cash Forecast"
    )
    _inject_sheet_commentary_cols(
        workbook, "Cash Flow Statement", commentary.get("cash_flow_statement") or {}, ws_name="Cash Flow Statement"
    )
    _inject_sheet_commentary_cols(
        workbook, "Q2 vs Budget", commentary.get("q2_vs_budget") or {}, ws_name="Q2 vs Budget"
    )
    _inject_sheet_commentary_cols(
        workbook, "GTM Review", commentary.get("gtm_review") or {}, ws_name="GTM Review"
    )
    if "GTM Review" in workbook.sheetnames:
        _apply_gtm_column_widths(workbook["GTM Review"])


def _apply_gtm_column_widths(ws) -> None:
    from openpyxl.utils import get_column_letter

    layout = SHEET_COMMENTARY_LAYOUT["GTM Review"]
    for letter, width in GTM_DATA_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for offset in range(3):
        col = layout["start_col"] + offset
        ws.column_dimensions[get_column_letter(col)].width = GTM_COMMENTARY_COL_WIDTH
    _extend_title_merge(ws, layout["start_col"] + 2)


def _sheet_commentary_layout(ws_name: str) -> dict[str, int]:
    if ws_name in SHEET_COMMENTARY_LAYOUT:
        return SHEET_COMMENTARY_LAYOUT[ws_name]
    return {"header_row": HEADER_ROW, "start_col": _sheet_commentary_start_col_legacy(ws)}


def _sheet_commentary_start_col_legacy(ws) -> int:
    """First commentary column = first column after the last populated header on row 2."""
    last = 1
    for col in range(1, ws.max_column + 5):
        val = ws.cell(HEADER_ROW, col).value
        if val is None:
            if col > last + 3:
                break
            continue
        if "commentary" in str(val).lower():
            return col
        last = col
    return last + 1


def _inject_sheet_commentary_cols(
    workbook: Any,
    sheet_name: str,
    row_commentary: dict[str, Any],
    *,
    ws_name: str | None = None,
) -> None:
    name = ws_name or sheet_name
    if name not in workbook.sheetnames or not row_commentary:
        return
    ws = workbook[name]
    layout = _sheet_commentary_layout(name)
    header_row = layout["header_row"]
    start_col = layout["start_col"]
    if ws.cell(header_row, start_col).value != COMMENTARY_HEADERS[0]:
        header_style = ws.cell(header_row, max(1, start_col - 1))
        if header_style.value is None:
            header_style = ws.cell(header_row, 1)
        for i, header in enumerate(COMMENTARY_HEADERS):
            row, col = _resolve_write_cell(ws, header_row, start_col + i)
            cell = ws.cell(row, col, header)
            _copy_cell_style(header_style, cell)
    label_to_row = _label_row_index(ws)
    for row_id, cols in row_commentary.items():
        excel_row = _row_for_commentary(row_id, label_to_row)
        if not excel_row:
            continue
        _apply_commentary_cell(ws, excel_row, start_col, cols.get("period_vs_budget", ""))
        _apply_commentary_cell(ws, excel_row, start_col + 1, cols.get("qtd_vs_budget", ""))
        _apply_commentary_cell(ws, excel_row, start_col + 2, cols.get("ytd_vs_budget", ""))
    if name in ("ARR Waterfall", "Income Statement", "Cash Forecast", "Cash Flow Statement"):
        _extend_title_merge(ws, start_col + 2)
    if name == "Q2 vs Budget":
        _extend_title_merge(ws, max(start_col + 2, 12))


def _label_row_index(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        for val in (a, b):
            if val is None:
                continue
            key = re.sub(r"^[\s+\-]+", "", str(val).strip().lower())
            if key:
                out[key] = r
    return out


def _row_for_commentary(row_id: str, label_to_row: dict[str, int]) -> int | None:
    direct = {
        "arr_new_business": "new business",
        "arr_expansion": "expansion",
        "arr_reactivation": "reactivation",
        "arr_contraction": "contraction",
        "arr_churn": "churn",
        "arr_net_new": "net new arr",
        "arr_ending": "ending arr",
        "is_total_revenue": "total revenue",
        "is_gross_profit": "gross profit",
        "is_sm": "sales & marketing",
        "is_rd": "research & development",
        "is_ga": "general & administrative",
        "is_ebitda": "ebitda",
        "gtm_organic_search": "organic search",
        "gtm_partner": "partner",
        "gtm_referral": "referral",
        "gtm_webinar": "webinar",
        "gtm_outbound": "outbound",
        "gtm_direct": "direct",
        "gtm_paid_search": "paid search",
        "gtm_paid_social": "paid social",
    }
    needle = direct.get(row_id)
    if needle and needle in label_to_row:
        return label_to_row[needle]
    if row_id.startswith("gtm_"):
        needle = row_id.replace("gtm_", "").replace("_", " ")
        return label_to_row.get(needle)
    return None


def build_mda_package_xlsx_bytes(
    *,
    template_path: Path,
    commentary: dict[str, Any],
    close_period: str,
    display: dict[str, Any] | None = None,
    bundle: ReportingBundle | None = None,
    deck: dict[str, Any] | None = None,
    ts_data: dict[str, Any] | None = None,
    cash_bridge_data: dict[str, Any] | None = None,
    channels: list[dict[str, Any]] | None = None,
) -> bytes:
    import openpyxl

    from app.services.reporting.export.mda_package_data import refresh_mda_workbook_data

    wb = openpyxl.load_workbook(template_path)
    if bundle is not None and deck is not None:
        refresh_mda_workbook_data(
            wb,
            bundle,
            deck,
            ts_data=ts_data,
            cash_bridge_data=cash_bridge_data,
            channels=channels,
        )
    inject_commentary(wb, commentary, close_period=close_period, display=display)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
