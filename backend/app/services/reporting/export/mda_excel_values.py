"""Numeric Excel cell values for SMPL_MDA_Package (formula-friendly $M units)."""

from __future__ import annotations

import re
from decimal import Decimal
from typing import Any

from app.services.reporting.export.schemas import ReportingBundle

ZERO = Decimal("0")

FMT_MONEY_M = r'\$#,##0.0;(\$#,##0.0);"-"'
FMT_VAR_M = r'\+$#,##0.0;(\$#,##0.0);"-"'
FMT_PCT = "0.0%"
FMT_HEADCOUNT = "0"
FMT_RATIO = "0.0"


def decimal_to_m(value: Decimal | float | int | None) -> float | None:
    if value is None:
        return None
    v = float(value)
    if v == 0:
        return 0.0
    return round(v / 1_000_000, 6)


def parse_deck_money_to_m(text: Any) -> float | None:
    """Parse fmt_deck_money strings into $M floats for Excel."""
    if text is None or text == "" or text == "—" or text == "n/a":
        return None
    if isinstance(text, (int, float)):
        return float(text)
    s = str(text).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    sign = -1 if s.startswith("-") or s.startswith("(") else 1
    s = s.lstrip("+-").rstrip(")").strip()
    mult = 1.0
    if s.upper().endswith("M"):
        mult = 1.0
        s = s[:-1]
    elif s.upper().endswith("K"):
        mult = 0.001
        s = s[:-1]
    try:
        return sign * round(float(s) * mult, 6)
    except ValueError:
        return None


def parse_deck_pct_to_ratio(text: Any) -> float | None:
    if text is None or text == "" or text == "—" or text == "n/a":
        return None
    if isinstance(text, (int, float)):
        v = float(text)
        return v / 100 if abs(v) > 1.5 else v
    s = str(text).strip().replace("%", "")
    try:
        v = float(s)
        return round(v / 100, 6)
    except ValueError:
        return None


def numeric_horizon_block(
    actual: Decimal,
    budget: Decimal,
    *,
    is_pct: bool = False,
) -> dict[str, float | None]:
    if is_pct:
        def _ratio(v: Decimal) -> float | None:
            if v is None:
                return None
            fv = float(v)
            return round(fv / 100, 6) if abs(fv) > 1.5 else round(fv, 6)

        act = _ratio(actual)
        bud = _ratio(budget)
        var = (act - bud) if act is not None and bud is not None else None
        var_pct = (var / bud) if var is not None and bud else None
        return {"actual": act, "budget": bud, "variance": var, "var_pct": var_pct}
    act_m = decimal_to_m(actual)
    bud_m = decimal_to_m(budget)
    var_m = (act_m - bud_m) if act_m is not None and bud_m is not None else None
    var_pct = (var_m / bud_m) if var_m is not None and bud_m else None
    return {"actual": act_m, "budget": bud_m, "variance": var_m, "var_pct": var_pct}


def numeric_block_from_formatted(block: dict[str, Any], *, is_pct: bool = False) -> dict[str, float | None]:
    if is_pct:
        act = parse_deck_pct_to_ratio(block.get("actual"))
        bud = parse_deck_pct_to_ratio(block.get("budget"))
        var = (act - bud) if act is not None and bud is not None else None
        var_pct = parse_deck_pct_to_ratio(block.get("var_pct"))
        if var_pct is None and var is not None and bud:
            var_pct = var / bud
        return {"actual": act, "budget": bud, "variance": var, "var_pct": var_pct}
    act = parse_deck_money_to_m(block.get("actual"))
    bud = parse_deck_money_to_m(block.get("budget"))
    var = parse_deck_money_to_m(block.get("variance") or block.get("var"))
    if var is None and act is not None and bud is not None:
        var = act - bud
    var_pct = parse_deck_pct_to_ratio(block.get("var_pct"))
    if var_pct is None and var is not None and bud:
        var_pct = var / bud
    return {"actual": act, "budget": bud, "variance": var, "var_pct": var_pct}


def budget_qtd_waterfall(bundle: ReportingBundle, wtype: str) -> Decimal:
    from app.services.reporting.export.board_platform_metrics import _wf
    from app.services.reporting.export.period_views import qtd_periods
    from app.services.reporting.period_utils import to_period

    total = ZERO
    for p in qtd_periods(bundle.as_of_period):
        total += _wf(bundle, "arr", wtype, to_period(p), "Budget")
    return total


def extend_title_merge(ws, end_col: int) -> None:
    from openpyxl.utils import get_column_letter

    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == 1 and merged.max_row == 1:
            ws.unmerge_cells(str(merged))
    ws.merge_cells(f"A1:{get_column_letter(end_col)}1")
