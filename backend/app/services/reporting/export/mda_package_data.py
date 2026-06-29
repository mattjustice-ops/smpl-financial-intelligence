"""Refresh SMPL_MDA_Package Excel tabs from board_platform_metrics / TS_DATA."""

from __future__ import annotations

import calendar
import re
from decimal import Decimal
from typing import Any

from app.services.reporting.export.board_period_ytd import _wf_by_period
from app.services.reporting.export.board_platform_metrics import headcount_totals
from app.services.reporting.export.effective_periods import export_fiscal_periods
from app.services.reporting.export.deck_payload_enriched import build_pl_detail_block
from app.services.reporting.export.mda_excel_values import (
    FMT_MONEY_M,
    FMT_PCT,
    FMT_VAR_M,
    budget_qtd_waterfall,
    decimal_to_m,
    extend_title_merge,
    numeric_block_from_formatted,
    numeric_horizon_block,
)
from app.services.reporting.export.schemas import ReportingBundle
from app.services.reporting.period_utils import to_period

ZERO = Decimal("0")
MONTH_COL = 2  # column B = Jan

try:
    from openpyxl.cell.cell import MergedCell
except ImportError:  # pragma: no cover
    MergedCell = type(None)  # type: ignore[misc,assignment]


def _resolve_write_cell(ws, row: int, col: int) -> tuple[int, int]:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return row, col
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col
    return row, col


def _set_cell(ws, row: int, col: int, value: Any, *, number_format: str | None = None) -> None:
    row, col = _resolve_write_cell(ws, row, col)
    cell = ws.cell(row, col, value)
    if number_format is not None:
        cell.number_format = number_format
SUBTOTAL_COLS = {
    "q1_actual": 14,
    "q2_actual": 15,
    "h1_actual": 16,
    "fy_actual": 17,
    "q1_budget": 18,
    "q2_budget": 19,
    "h1_budget": 20,
    "fy_budget": 21,
    "q1_var": 22,
    "q2_var": 23,
    "h1_var": 24,
    "fy_var": 25,
}


def _norm_label(text: str) -> str:
    return re.sub(r"^[\s+\-]+", "", str(text or "").strip().lower())


def _to_m(val: Decimal | float | int | None) -> float | None:
    if val is None:
        return None
    v = float(val)
    if v == 0:
        return 0.0
    return round(v / 1_000_000, 6)


def _pct_decimal(numerator: Decimal, denominator: Decimal) -> float | None:
    if not denominator:
        return None
    return round(float(numerator / denominator), 4)


def _periods(bundle: ReportingBundle) -> list[str]:
    return [to_period(p) for p in export_fiscal_periods(bundle.start_period, bundle.end_period)]


def _scenario_values(
    bundle: ReportingBundle,
    *,
    key: str,
    wtype: str,
    as_of: str,
    negate: bool = False,
) -> tuple[list[float | None], list[float | None]]:
    """Monthly Actual/Forecast (cols B–M) and parallel Budget series."""
    actual = _wf_by_period(bundle, key, wtype, "Actual")
    forecast = _wf_by_period(bundle, key, wtype, "Forecast")
    budget = _wf_by_period(bundle, key, wtype, "Budget")
    act_out: list[float | None] = []
    bud_out: list[float | None] = []
    for period in _periods(bundle):
        p = to_period(period)
        if p <= as_of:
            raw = actual.get(p, ZERO)
        else:
            raw = forecast.get(p, ZERO)
        if negate and raw:
            raw = -abs(raw)
        bud_raw = budget.get(p, ZERO)
        if negate and bud_raw:
            bud_raw = -abs(bud_raw)
        act_out.append(_to_m(raw))
        bud_out.append(_to_m(bud_raw))
    return act_out, bud_out


def _ts_monthly(
    ts_data: dict[str, Any] | None,
    field: str,
    periods: list[str],
    as_of: str,
) -> tuple[list[float | None], list[float | None]]:
    if not ts_data:
        return [None] * len(periods), [None] * len(periods)
    act_blk = ts_data.get("Actual") or {}
    fc_blk = ts_data.get("Forecast") or {}
    bud_blk = ts_data.get("Budget") or {}
    act_is = act_blk.get("is") or {}
    fc_is = fc_blk.get("is") or {}
    bud_is = bud_blk.get("is") or {}
    act_out: list[float | None] = []
    bud_out: list[float | None] = []
    for period in periods:
        p = to_period(period)
        if p <= as_of:
            row = act_is.get(p) or {}
        else:
            row = fc_is.get(p) or {}
        bud_row = bud_is.get(p) or {}
        act_out.append(_to_m(row.get(field)))
        bud_out.append(_to_m(bud_row.get(field)))
    return act_out, bud_out


def _ts_cfs_monthly(
    ts_data: dict[str, Any] | None,
    field: str,
    periods: list[str],
    as_of: str,
) -> tuple[list[float | None], list[float | None]]:
    if not ts_data:
        return [None] * len(periods), [None] * len(periods)
    act_blk = ts_data.get("Actual") or {}
    fc_blk = ts_data.get("Forecast") or {}
    bud_blk = ts_data.get("Budget") or {}
    act_cfs = act_blk.get("cfs") or {}
    fc_cfs = fc_blk.get("cfs") or {}
    bud_cfs = bud_blk.get("cfs") or {}
    act_out: list[float | None] = []
    bud_out: list[float | None] = []
    for period in periods:
        p = to_period(period)
        if p <= as_of:
            row = act_cfs.get(p) or {}
        else:
            row = fc_cfs.get(p) or {}
        bud_row = bud_cfs.get(p) or {}
        act_out.append(_to_m(row.get(field)))
        bud_out.append(_to_m(bud_row.get(field)))
    return act_out, bud_out


def _bridge_monthly(
    cash_bridge_data: dict[str, Any] | None,
    field: str,
    periods: list[str],
    as_of: str,
) -> tuple[list[float | None], list[float | None]]:
    if not cash_bridge_data:
        return [None] * len(periods), [None] * len(periods)
    act = cash_bridge_data.get("Actual") or {}
    bud = cash_bridge_data.get("Budget") or {}
    fc = cash_bridge_data.get("Forecast") or {}
    act_out: list[float | None] = []
    bud_out: list[float | None] = []
    for period in periods:
        p = to_period(period)
        if p <= as_of:
            row = act.get(p) or {}
        else:
            row = fc.get(p) or {}
        bud_row = bud.get(p) or {}
        act_out.append(_to_m(row.get(field)))
        bud_out.append(_to_m(bud_row.get(field)))
    return act_out, bud_out


def _write_monthly_row(
    ws,
    row: int,
    act_vals: list[float | None],
    bud_vals: list[float | None],
    *,
    as_of: str,
    periods: list[str],
    is_flow: bool = True,
    is_pct: bool = False,
) -> None:
    """Write Jan–Dec then Q/H/FY subtotals."""
    for i, val in enumerate(act_vals):
        if val is not None:
            fmt = FMT_PCT if is_pct else FMT_MONEY_M
            _set_cell(ws, row, MONTH_COL + i, val, number_format=fmt)
    q1a = q2a = h1a = fya = None
    q1b = q2b = h1b = fyb = None
    if is_flow:
        q1a = sum(v or 0 for v in act_vals[0:3]) or None
        q2a = sum(v or 0 for v in act_vals[3:6]) or None
        h1a = sum(v or 0 for v in act_vals[0:6]) or None
        fya = sum(v or 0 for v in act_vals) or None
        q1b = sum(v or 0 for v in bud_vals[0:3]) or None
        q2b = sum(v or 0 for v in bud_vals[3:6]) or None
        h1b = sum(v or 0 for v in bud_vals[0:6]) or None
        fyb = sum(v or 0 for v in bud_vals) or None
    else:
        q1a = act_vals[2]
        q2a = act_vals[5]
        h1a = act_vals[5]
        fya = act_vals[11] if len(act_vals) > 11 else act_vals[-1]
        q1b = bud_vals[2]
        q2b = bud_vals[5]
        h1b = bud_vals[5]
        fyb = bud_vals[11] if len(bud_vals) > 11 else bud_vals[-1]

    def _var(a: float | None, b: float | None) -> tuple[float | None, float | None]:
        if a is None or b is None:
            return None, None
        diff = a - b
        pct = (diff / b) if b else None
        return diff, pct

    pairs = (
        (SUBTOTAL_COLS["q1_actual"], q1a, q1b),
        (SUBTOTAL_COLS["q2_actual"], q2a, q2b),
        (SUBTOTAL_COLS["h1_actual"], h1a, h1b),
        (SUBTOTAL_COLS["fy_actual"], fya, fyb),
        (SUBTOTAL_COLS["q1_budget"], None, q1b),
        (SUBTOTAL_COLS["q2_budget"], None, q2b),
        (SUBTOTAL_COLS["h1_budget"], None, h1b),
        (SUBTOTAL_COLS["fy_budget"], None, fyb),
    )
    for col, a, b in pairs:
        if col in (SUBTOTAL_COLS["q1_budget"], SUBTOTAL_COLS["q2_budget"], SUBTOTAL_COLS["h1_budget"], SUBTOTAL_COLS["fy_budget"]):
            if b is not None:
                _set_cell(ws, row, col, b, number_format=FMT_MONEY_M)
        elif a is not None:
            _set_cell(ws, row, col, a, number_format=FMT_MONEY_M)
    for label, a, b in (("q1", q1a, q1b), ("q2", q2a, q2b), ("h1", h1a, h1b), ("fy", fya, fyb)):
        diff, pct = _var(a, b)
        if diff is not None:
            _set_cell(ws, row, SUBTOTAL_COLS[f"{label}_var"], diff, number_format=FMT_VAR_M)
        if pct is not None and not is_pct:
            _set_cell(ws, row, SUBTOTAL_COLS[f"{label}_var"] + 4, pct, number_format=FMT_PCT)


def _label_row_map(ws, col: int = 1) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, col).value
        if val:
            out[_norm_label(str(val))] = r
    return out


def _find_row(labels: dict[str, int], *needles: str) -> int | None:
    for needle in needles:
        n = _norm_label(needle)
        if n in labels:
            return labels[n]
        for key, row in labels.items():
            if n in key:
                return row
    return None


def refresh_arr_waterfall_sheet(ws, bundle: ReportingBundle, *, as_of: str, month_label: str, year: str) -> None:
    _set_cell(ws, 1, 1, f"SMPL · ARR / MRR Waterfall · Q{_quarter_num(as_of)} {year} Close — {month_label} Actuals")
    periods = _periods(bundle)
    labels = _label_row_map(ws)
    specs = (
        ("beginning arr", "beginning", False, False),
        ("new business", "new_business", False, True),
        ("expansion", "expansion", False, True),
        ("reactivation", "reactivation", False, True),
        ("contraction", "contraction", True, True),
        ("churn", "churn", True, True),
        ("net new arr", "new_arr", False, True),
        ("ending arr", "ending_arr", False, False),
    )
    for needle, wtype, negate, is_flow in specs:
        row = _find_row(labels, needle, f"+ {needle}", f"- {needle}")
        if not row:
            continue
        act, bud = _scenario_values(bundle, key="arr", wtype=wtype, as_of=as_of, negate=negate)
        if wtype == "new_arr" and all(v in (None, 0.0) for v in act):
            act, bud = _scenario_values(bundle, key="arr", wtype="new_business", as_of=as_of)
        if wtype == "ending_arr":
            for alt in ("ending",):
                if all(v in (None, 0.0) for v in act):
                    act, bud = _scenario_values(bundle, key="arr", wtype=alt, as_of=as_of)
        _write_monthly_row(ws, row, act, bud, as_of=as_of, periods=periods, is_flow=is_flow)


def refresh_income_statement_sheet(
    ws,
    bundle: ReportingBundle,
    ts_data: dict[str, Any] | None,
    *,
    as_of: str,
    year: str,
) -> None:
    _set_cell(ws, 1, 1, f"SMPL · Income Statement · Jan–Dec {year}")
    periods = _periods(bundle)
    labels = _label_row_map(ws)
    line_specs = (
        ("total revenue", "revenue", False),
        ("cost of revenue", "cogs", False),
        ("gross profit", "gross_profit", False),
        ("sales & marketing", "sm", False),
        ("research & development", "rd", False),
        ("general & administrative", "ga", False),
        ("total opex", "total_opex", False),
        ("ebitda", "ebitda", False),
        ("depreciation & amort", "da", False),
        ("operating income", "op_income", False),
    )
    revenue_act: list[float | None] = []
    gp_act: list[float | None] = []
    for needle, field, is_pct in line_specs:
        row = _find_row(labels, needle)
        if not row:
            continue
        act, bud = _ts_monthly(ts_data, field, periods, as_of)
        _write_monthly_row(ws, row, act, bud, as_of=as_of, periods=periods, is_flow=True, is_pct=is_pct)
        if field == "revenue":
            revenue_act = act
        if field == "gross_profit":
            gp_act = act
    gm_row = _find_row(labels, "gross margin %")
    if gm_row and revenue_act and gp_act:
        gm_vals = []
        for rev, gp in zip(revenue_act, gp_act):
            if rev and gp is not None:
                gm_vals.append(round(gp / rev, 4) if rev else None)
            else:
                gm_vals.append(None)
        _write_monthly_row(ws, gm_row, gm_vals, gm_vals, as_of=as_of, periods=periods, is_flow=False, is_pct=True)


def refresh_cash_forecast_sheet(
    ws,
    cash_bridge_data: dict[str, Any] | None,
    *,
    as_of: str,
    year: str,
) -> None:
    _set_cell(ws, 1, 1, f"SMPL · Cash Flow Bridge · Jan–Dec {year}")
    periods = [f"{year}-{m:02d}" for m in range(1, 13)]
    labels = _label_row_map(ws)
    specs = (
        ("beginning cash", "beginning_cash", False),
        ("collections", "collections", True),
        ("payroll", "payroll", True),
        ("commissions", "commission", True),
        ("vendor", "vendor", True),
        ("tax", "tax", True),
        ("interest", "interest", True),
        ("capital expenditures", "capex", True),
        ("ending cash balance", "ending_cash", False),
    )
    for needle, field, is_flow in specs:
        row = _find_row(labels, needle, needle.replace("commission", "commissions"))
        if not row:
            continue
        act, bud = _bridge_monthly(cash_bridge_data, field, periods, as_of)
        _write_monthly_row(ws, row, act, bud, as_of=as_of, periods=periods, is_flow=is_flow)


def refresh_cash_flow_statement_sheet(
    ws,
    ts_data: dict[str, Any] | None,
    *,
    as_of: str,
    year: str,
) -> None:
    _set_cell(ws, 1, 1, f"SMPL · Cash Flow Statement (GAAP Indirect Method) · Jan–Dec {year}")
    periods = [f"{year}-{m:02d}" for m in range(1, 13)]
    labels = _label_row_map(ws)
    specs = (
        ("net income", "net_income", True),
        ("depreciation & amort", "da", True),
        ("change in ar", "chg_ar", True),
        ("change in deferred revenue", "chg_dr", True),
        ("change in ap", "chg_ap", True),
        ("cash from operations", "cfo", True),
        ("capital expenditures", "capex", True),
        ("cash from investing", "cfi", True),
        ("cash from financing", "cff", True),
        ("net change in cash", "net_change", True),
        ("beginning cash balance", "beginning_cash", False),
        ("ending cash balance", "ending_cash", False),
    )
    for needle, field, is_flow in specs:
        row = _find_row(labels, needle)
        if not row:
            continue
        act, bud = _ts_cfs_monthly(ts_data, field, periods, as_of)
        _write_monthly_row(ws, row, act, bud, as_of=as_of, periods=periods, is_flow=is_flow)


def _write_q2_ytd_row(
    ws,
    row: int,
    qtd: dict[str, float | None],
    ytd: dict[str, float | None],
    *,
    is_pct: bool = False,
) -> None:
    money_fmt = FMT_PCT if is_pct else FMT_MONEY_M
    var_fmt = FMT_PCT if is_pct else FMT_VAR_M
    for offset, blk in enumerate((qtd, ytd)):
        base = 2 + offset * 4
        _set_cell(ws, row, base, blk.get("actual"), number_format=money_fmt)
        _set_cell(ws, row, base + 1, blk.get("budget"), number_format=money_fmt)
        _set_cell(ws, row, base + 2, blk.get("variance"), number_format=var_fmt)
        _set_cell(ws, row, base + 3, blk.get("var_pct"), number_format=FMT_PCT)


def refresh_q2_vs_budget_sheet(
    ws,
    deck: dict[str, Any],
    bundle: ReportingBundle,
    *,
    ts_data: dict[str, Any] | None,
    as_of: str,
    year: str,
) -> None:
    from app.services.reporting.export.board_period_ytd import rollup_waterfall_metric
    from app.services.reporting.export.board_platform_metrics import (
        ending_arr_at_period,
        gross_margin_pct_horizon,
        rollup_cash,
        rollup_ebitda,
        rollup_revenue,
        _budget_qtd_flow,
        _cash_wf_amount,
    )
    from app.services.reporting.export.period_views import qtd_periods, ytd_periods

    month = calendar.month_name[int(as_of[5:7])]
    q = _quarter_num(as_of)
    _set_cell(ws, 1, 1, f"SMPL · Q{q} {year} Variance Analysis — Actuals vs Budget")
    _set_cell(ws, 2, 1, f"Q{q} {year} (Apr–Jun) and YTD (Jan–{month}) · Prepared by Finance · Confidential")
    pl_detail = build_pl_detail_block(bundle, ts_data)
    labels = _label_row_map(ws)
    qtd_p = qtd_periods(as_of)
    ytd_p = ytd_periods(as_of)

    rev = rollup_revenue(bundle)
    ebitda = rollup_ebitda(bundle)
    cash = rollup_cash(bundle)
    qtd_rev_bud = _budget_qtd_flow(bundle, "revenue")
    qtd_ebitda_bud = _budget_qtd_flow(bundle, "ebitda")

    is_rows = (
        (
            "total revenue",
            numeric_block_from_formatted((pl_detail.get("qtd") or {}).get("revenue") or {}),
            numeric_block_from_formatted((pl_detail.get("ytd") or {}).get("revenue") or {}),
        ),
        (
            "gross profit",
            numeric_block_from_formatted((pl_detail.get("qtd") or {}).get("gross_profit") or {}),
            numeric_block_from_formatted((pl_detail.get("ytd") or {}).get("gross_profit") or {}),
        ),
        (
            "gross margin %",
            numeric_horizon_block(
                gross_margin_pct_horizon(bundle, qtd_p, "Actual") or ZERO,
                gross_margin_pct_horizon(bundle, qtd_p, "Budget") or ZERO,
                is_pct=True,
            ),
            numeric_horizon_block(
                gross_margin_pct_horizon(bundle, ytd_p, "Actual") or ZERO,
                gross_margin_pct_horizon(bundle, ytd_p, "Budget") or ZERO,
                is_pct=True,
            ),
        ),
        ("sales & marketing", numeric_block_from_formatted((pl_detail.get("qtd") or {}).get("sm") or {}), numeric_block_from_formatted((pl_detail.get("ytd") or {}).get("sm") or {})),
        ("research & development", numeric_block_from_formatted((pl_detail.get("qtd") or {}).get("rd") or {}), numeric_block_from_formatted((pl_detail.get("ytd") or {}).get("rd") or {})),
        ("general & administrative", numeric_block_from_formatted((pl_detail.get("qtd") or {}).get("ga") or {}), numeric_block_from_formatted((pl_detail.get("ytd") or {}).get("ga") or {})),
        ("ebitda", numeric_horizon_block(ebitda.qtd, qtd_ebitda_bud), numeric_horizon_block(ebitda.ytd, ebitda.budget_ytd)),
    )
    for needle, qtd_blk, ytd_blk in is_rows:
        row = _find_row(labels, needle)
        if row:
            _write_q2_ytd_row(ws, row, qtd_blk, ytd_blk, is_pct="margin" in needle)

    arr_act = ending_arr_at_period(bundle, as_of, "Actual")
    arr_bud = ending_arr_at_period(bundle, as_of, "Budget")
    arr_row = _find_row(labels, "ending arr (jun eop)", "ending arr")
    if arr_row:
        eop = numeric_horizon_block(arr_act, arr_bud)
        _write_q2_ytd_row(ws, arr_row, eop, eop)

    cash_act = _cash_wf_amount(bundle, as_of, "Actual", "ending_cash", "ending") or cash.current_month
    cash_bud = _cash_wf_amount(bundle, as_of, "Budget", "ending_cash", "ending") or cash.budget_cm
    cash_row = _find_row(labels, "ending cash (jun eop)", "ending cash")
    if cash_row:
        eop = numeric_horizon_block(cash_act, cash_bud)
        _write_q2_ytd_row(ws, cash_row, eop, eop)

    cfo_row = _find_row(labels, "q2 / ytd cfo", "cfo")
    if cfo_row:
        from app.services.reporting.export.is_line_resolver import fs_sum_periods

        qtd_cfo = numeric_horizon_block(
            fs_sum_periods(bundle, qtd_p, "Actual", "cfo"),
            fs_sum_periods(bundle, qtd_p, "Budget", "cfo"),
        )
        ytd_cfo = numeric_horizon_block(
            fs_sum_periods(bundle, ytd_p, "Actual", "cfo"),
            fs_sum_periods(bundle, ytd_p, "Budget", "cfo"),
        )
        _write_q2_ytd_row(ws, cfo_row, qtd_cfo, ytd_cfo)

    arr_specs = (
        ("new business", "new_business"),
        ("expansion", "expansion"),
        ("reactivation", "reactivation"),
        ("contraction", "contraction"),
        ("churn", "churn"),
        ("net new arr", "new_arr"),
    )
    for needle, wtype in arr_specs:
        row = _find_row(labels, needle)
        if not row:
            continue
        rollup = rollup_waterfall_metric(bundle, "arr", wtype, flow=True)
        qtd_bud = budget_qtd_waterfall(bundle, wtype)
        if wtype == "new_arr" and qtd_bud == ZERO:
            qtd_bud = budget_qtd_waterfall(bundle, "new_business")
        qtd = numeric_horizon_block(rollup.qtd, qtd_bud)
        ytd = numeric_horizon_block(rollup.ytd, rollup.budget_ytd)
        _write_q2_ytd_row(ws, row, qtd, ytd)

    extend_title_merge(ws, 12)


def refresh_gtm_review_sheet(ws, channels: list[dict[str, Any]], *, year: str) -> None:
    _set_cell(ws, 1, 1, f"SMPL · GTM & Marketing Channel Efficiency · YTD {year}")
    headers = ["Channel", "Mktg Spend ($M)", "Pipeline Created ($M)", "Closed Won ($M)", "Pipeline / $1 Spend", "Efficiency Rank"]
    for c, h in enumerate(headers, start=1):
        _set_cell(ws, 2, c, h)
    for i, ch in enumerate(channels[:12], start=3):
        _set_cell(ws, i, 1, ch.get("name", ""))
        _set_cell(ws, i, 2, round((ch.get("spend_raw") or 0) / 1_000_000, 2))
        _set_cell(ws, i, 3, round((ch.get("pipeline_raw") or 0) / 1_000_000, 2))
        closed = ch.get("closed_won_raw")
        if closed is not None:
            _set_cell(ws, i, 4, round(float(closed) / 1_000_000, 2))
        _set_cell(ws, i, 5, ch.get("efficiency_x", ""))
        _set_cell(ws, i, 6, i - 2)


def refresh_headcount_sheet(ws, bundle: ReportingBundle, *, as_of: str, year: str) -> None:
    _set_cell(ws, 1, 1, f"SMPL · Headcount & Productivity · Jan–{calendar.month_name[int(as_of[5:7])]} {year}")
    close_month = int(as_of[5:7])
    month_cols = {m: 2 + (m - 1) for m in range(1, close_month + 1)}
    dept_rows: dict[str, dict[int, int]] = {}
    for row in bundle.headcount:
        if row.scenario != "Actual":
            continue
        dept = str(row.department or "Other")
        p = to_period(str(row.period))
        try:
            m = int(p[5:7])
        except ValueError:
            continue
        dept_rows.setdefault(dept, {})[m] = dept_rows.get(dept, {}).get(m, 0) + int(Decimal(str(row.headcount or 0)))
    labels = _label_row_map(ws)
    template_depts = (
        "sales",
        "engineering",
        "customer success",
        "marketing",
        "g&a + support",
        "product",
        "total",
    )

    def _match_dept(name: str) -> dict[int, int] | None:
        n = _norm_label(name)
        for dept, months in dept_rows.items():
            dn = _norm_label(dept)
            if n in dn or dn in n or (n == "g&a + support" and ("g&a" in dn or "general" in dn)):
                return months
        return None

    total_by_month: dict[int, int] = {m: 0 for m in range(1, close_month + 1)}
    for r in range(3, 10):
        dept_label = ws.cell(r, 1).value
        if not dept_label:
            continue
        months = _match_dept(str(dept_label))
        if str(dept_label).strip().lower() == "total":
            months = total_by_month
        elif months:
            for m, v in months.items():
                total_by_month[m] = total_by_month.get(m, 0) + v
        if not months:
            continue
        for m, col in month_cols.items():
            _set_cell(ws, r, col, months.get(m, 0))
        act = months.get(close_month, 0)
        bud_months = _dept_hc_by_month(bundle, str(dept_label), "Budget")
        bud = bud_months.get(close_month, 0) if bud_months else 0
        if str(dept_label).strip().lower() == "total":
            bud, _ = headcount_totals(bundle, as_of, "Budget")
        _set_cell(ws, r, 8, bud)
        _set_cell(ws, r, 9, act - bud if act or bud else 0)


def _dept_hc_by_month(bundle: ReportingBundle, dept_label: str, scenario: str) -> dict[int, int] | None:
    n = _norm_label(dept_label)
    if n == "total":
        return None
    out: dict[int, int] = {}
    for row in bundle.headcount:
        if row.scenario != scenario:
            continue
        dn = _norm_label(str(row.department or ""))
        if n not in dn and dn not in n and not (n == "g&a + support" and ("g&a" in dn or "general" in dn)):
            continue
        try:
            m = int(to_period(str(row.period))[5:7])
        except ValueError:
            continue
        out[m] = out.get(m, 0) + int(Decimal(str(row.headcount or 0)))
    return out or None


def _quarter_num(period: str) -> int:
    return (int(period[5:7]) - 1) // 3 + 1


def refresh_mda_workbook_data(
    workbook: Any,
    bundle: ReportingBundle,
    deck: dict[str, Any],
    *,
    ts_data: dict[str, Any] | None = None,
    cash_bridge_data: dict[str, Any] | None = None,
    channels: list[dict[str, Any]] | None = None,
) -> None:
    """Overwrite data tabs from warehouse-aligned deck payload."""
    as_of = bundle.as_of_period
    year = as_of[:4]
    month_label = calendar.month_name[int(as_of[5:7])]
    if "ARR Waterfall" in workbook.sheetnames:
        refresh_arr_waterfall_sheet(workbook["ARR Waterfall"], bundle, as_of=as_of, month_label=month_label, year=year)
    if "Income Statement" in workbook.sheetnames:
        refresh_income_statement_sheet(workbook["Income Statement"], bundle, ts_data, as_of=as_of, year=year)
    if "Cash Forecast" in workbook.sheetnames:
        refresh_cash_forecast_sheet(workbook["Cash Forecast"], cash_bridge_data, as_of=as_of, year=year)
    if "Cash Flow Statement" in workbook.sheetnames:
        refresh_cash_flow_statement_sheet(workbook["Cash Flow Statement"], ts_data, as_of=as_of, year=year)
    if "Q2 vs Budget" in workbook.sheetnames:
        refresh_q2_vs_budget_sheet(
            workbook["Q2 vs Budget"],
            deck,
            bundle,
            ts_data=ts_data,
            as_of=as_of,
            year=year,
        )
    if "GTM Review" in workbook.sheetnames and channels:
        refresh_gtm_review_sheet(workbook["GTM Review"], channels, year=year)
    if "Headcount" in workbook.sheetnames:
        refresh_headcount_sheet(workbook["Headcount"], bundle, as_of=as_of, year=year)
