"""Tests for Prompt 2 MDA package payload and injection helpers."""

from __future__ import annotations

import io

from app.services.dashboard.schemas import ExecutiveFlowResponse
from app.services.reporting.export.mda_package_injection import (
    CHAR_LIMITS,
    resolve_mda_package_template,
    trim_commentary_response,
)
from app.services.reporting.export.mda_package_payload import build_mda_package_payload
from app.services.reporting.export.schemas import ExportValidationSummary, ReportingBundle


def _minimal_bundle(**overrides) -> ReportingBundle:
    base = dict(
        organization_id="8571e520-0687-4516-bdee-379f37c58c1f",
        organization_name="SMPL.ai",
        currency="USD",
        scenario="Combined",
        start_period="2026-01",
        end_period="2026-12",
        as_of_period="2026-06",
        period_label="June 2026",
        executive_flow=ExecutiveFlowResponse(
            organization_id="8571e520-0687-4516-bdee-379f37c58c1f",
            scenario="Combined",
            start_period="2026-01",
            end_period="2026-12",
            as_of_period="2026-06",
        ),
        comparison_waterfalls={},
        validation=ExportValidationSummary(status="pass"),
    )
    base.update(overrides)
    return ReportingBundle(**base)


def test_resolve_mda_package_template_finds_repo_copy() -> None:
    path = resolve_mda_package_template()
    assert path is not None
    assert path.name.endswith(".xlsx")


def test_trim_commentary_enforces_limits() -> None:
    long_text = "A" * 500
    raw = {
        "variance_commentary": {
            "vc_arr_ending": {
                "period_vs_budget": long_text,
                "qtd_vs_budget": "Short QTD line.",
                "ytd_vs_budget": "Short YTD line.",
            }
        }
    }
    trimmed = trim_commentary_response(raw)
    limit = CHAR_LIMITS["variance_commentary"]["period_vs_budget"]
    assert len(trimmed["variance_commentary"]["vc_arr_ending"]["period_vs_budget"]) <= limit


def test_build_mda_package_payload_shape() -> None:
    payload = build_mda_package_payload(_minimal_bundle())
    assert payload["task"] == "mda_full_package"
    assert "variance_commentary" in payload["sheets"]
    vc_rows = payload["sheets"]["variance_commentary"]["rows"]
    assert any(r["row_id"] == "vc_arr_ending" for r in vc_rows)
    display = payload["variance_commentary_display"]
    assert len(display["rows"]) == len(vc_rows)
    assert payload["payload_meta"]["data_map"]["canonical_engine"] == "board_platform_metrics"
    assert not payload["tie_out_warnings"]


def test_deliverable_data_map_in_deck_payload() -> None:
    from app.services.reporting.export.board_platform_metrics import build_deck_payload

    deck = build_deck_payload(_minimal_bundle())
    assert deck["data_map"]["shared_payload_builder"] == "build_deck_payload"
    assert "variance_commentary_rows" in deck["data_map"]


def test_refresh_mda_workbook_data_on_template() -> None:
    from app.services.reporting.export.board_platform_metrics import build_deck_payload
    from app.services.reporting.export.mda_package_data import refresh_mda_workbook_data

    template = resolve_mda_package_template()
    assert template is not None
    import openpyxl

    bundle = _minimal_bundle()
    deck = build_deck_payload(bundle)
    wb = openpyxl.load_workbook(template)
    refresh_mda_workbook_data(wb, bundle, deck)
    assert "June" in str(wb["ARR Waterfall"].cell(1, 1).value)
    assert "SMPL" in str(wb["Income Statement"].cell(1, 1).value)


def test_inject_commentary_into_template() -> None:
    from app.services.reporting.export.mda_package_injection import (
        build_mda_package_xlsx_bytes,
        _detect_variance_layout,
    )

    template = resolve_mda_package_template()
    assert template is not None
    import openpyxl

    wb = openpyxl.load_workbook(template)
    layout = _detect_variance_layout(wb["Variance Commentary"])
    assert layout.metric_start_col == 3
    assert layout.commentary_start_col >= 12
    commentary = {
        "variance_commentary": {
            "vc_arr_ending": {
                "period_vs_budget": "June ARR test period line.",
                "qtd_vs_budget": "Q2 ARR test qtd line.",
                "ytd_vs_budget": "YTD ARR test ytd line.",
            }
        },
        "risks_and_opportunities": {
            "ro_gtm_paid_search": {
                "description": "Paid search efficiency risk test.",
                "recommended_action": "Reallocate budget to Partner.",
            }
        },
    }
    display = {
        "sheet_title": "SMPL · Variance Commentary · June 2026",
        "rows": [
            {
                "row_id": "vc_arr_ending",
                "category": "ARR",
                "metric": "Ending ARR",
                "value_kind": "money",
                "cm": {"actual": "$86.10M", "budget": "$84.89M", "var": "$+1.21M"},
                "qtd": {"actual": "$86.10M", "budget": "$84.89M", "var": "$+1.21M"},
                "ytd": {"actual": "$86.10M", "budget": "$84.89M", "var": "$+1.21M"},
            },
            {
                "row_id": "vc_sm_pipeline",
                "category": "GTM",
                "metric": "S&M / Pipeline",
                "value_kind": "money",
                "cm": {"actual": "$3.15M", "budget": "$2.65M", "var": "$+0.50M"},
                "qtd": {"actual": "$9.29M", "budget": "$7.59M", "var": "$+1.70M"},
                "ytd": {"actual": "$18.29M", "budget": "$14.22M", "var": "$+4.07M"},
            },
            {
                "row_id": "vc_cash",
                "category": "Cash",
                "metric": "Ending Cash",
                "value_kind": "money",
                "cm": {"actual": "$70.61M", "budget": "$48.36M", "var": "$+22.25M"},
                "qtd": {"actual": "$70.61M", "budget": "$48.36M", "var": "$+22.25M"},
                "ytd": {"actual": "$70.61M", "budget": "$48.36M", "var": "$+22.25M"},
            },
            {
                "row_id": "vc_headcount",
                "category": "Headcount",
                "metric": "Total HC",
                "value_kind": "headcount",
                "cm": {"actual": "137", "budget": "137", "var": "—"},
                "qtd": {"actual": "137", "budget": "137", "var": "—"},
                "ytd": {"actual": "137", "budget": "137", "var": "—"},
            },
        ],
    }
    xlsx = build_mda_package_xlsx_bytes(
        template_path=template,
        commentary=commentary,
        close_period="2026-06",
        display=display,
    )
    assert len(xlsx) > 10_000
    out_wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = out_wb["Variance Commentary"]
    assert ws.cell(8, 4).value == 2.65
    assert ws.cell(9, 9).value == 70.61
    assert ws.cell(9, 10).value == 48.36
    assert ws.cell(10, 3).value == 137
    assert layout.commentary_start_col >= 12
    assert "Period vs Budget Commentary" in str(ws.cell(2, layout.commentary_start_col).value)
