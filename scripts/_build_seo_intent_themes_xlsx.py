"""Build SEO_Keyword_Strategy_Intent_Themes.xlsx for Matt review."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "marketing" / "SEO_Keyword_Strategy_Intent_Themes.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
P0_FILL = PatternFill("solid", fgColor="FEF3C7")
BRAND_FILL = PatternFill("solid", fgColor="E0E7FF")
NEG_FILL = PatternFill("solid", fgColor="FEE2E2")


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, max_width: int = 56) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 12
        for cell in col:
            if cell.value:
                width = min(max_width, max(width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_table(ws, headers: list[str], rows: list[list], freeze: str = "A2") -> None:
    ws.append(headers)
    style_header(ws)
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)
    ws.freeze_panes = freeze
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def highlight_priority(ws, priority_col: int, brand_values: set[str] | None = None) -> None:
    brand_values = brand_values or set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        pri = str(row[priority_col - 1].value or "")
        if pri == "P0":
            row[priority_col - 1].fill = P0_FILL
        if pri.lower().startswith("brand"):
            row[priority_col - 1].fill = BRAND_FILL
        # theme name / term in col 1
        if str(row[0].value or "") in brand_values:
            row[0].fill = BRAND_FILL


THEMES = [
    # theme, intent, priority, primary URL, url status, notes
    [
        "SaaS Financial Intelligence",
        "Commercial",
        "P0",
        "/",
        "Live",
        "Category umbrella; one-liner product theme. Align homepage/demo/pricing language.",
    ],
    [
        "SaaS FP&A Software",
        "Transactional / Commercial",
        "P0",
        "/",
        "Live",
        "Primary organic + ads north star with board/ARR. Titles already pivoted from AI OS.",
    ],
    [
        "ARR Reporting",
        "Commercial / Informational",
        "P0",
        "/blog/arr-waterfall-vs-gaap-revenue",
        "Live",
        "One-liner theme. Pair waterfall + governance + planned billing-vs-CRM.",
    ],
    [
        "Board Reporting",
        "Commercial / Problem",
        "P0",
        "/blog/saas-board-reporting-arr-cash-pl",
        "Live",
        "One-liner theme. Also homepage story; strong ads match.",
    ],
    [
        "Executive Reporting",
        "Commercial",
        "P1",
        "/",
        "Live (implicit)",
        "Ads-leaning cousin of board reporting; avoid pure 'executive dashboard' BI intent.",
    ],
    [
        "Cash Forecasting",
        "Informational / Commercial",
        "P1",
        "/blog/saas-cash-forecast",
        "Planned",
        "One-liner. Bridge from live board ARR–cash–P&L post; soft demo CTA.",
    ],
    [
        "Scenario Analysis",
        "Commercial",
        "P1",
        "/",
        "Planned / thin",
        "One-liner (Scenario Planning / Scenario Analysis). Supporting ads + future SEO page.",
    ],
    [
        "AI CFO Copilot",
        "Commercial",
        "P0",
        "/",
        "Live (implicit) — dedicated landing planned",
        "One-liner. Strong ads theme; SEO needs clearer landing for message match.",
    ],
    [
        "Pipeline Reporting",
        "Commercial",
        "P2",
        "/",
        "Live (implicit)",
        "Supporting SaaS operating-model language; secondary keyword theme.",
    ],
    [
        "Revenue Forecasting",
        "Commercial",
        "P1",
        "/",
        "Live (implicit)",
        "Pair with cash forecast and scenario themes.",
    ],
    [
        "SaaS KPI Reporting",
        "Commercial",
        "P2",
        "/glossary",
        "Live (hub)",
        "Supporting; glossary + board pack internal links.",
    ],
    [
        "Variance Analysis",
        "Informational / Commercial",
        "P1",
        "/blog",
        "Live (related posts)",
        "MD&A / commentary cluster; SEO literacy then ads.",
    ],
    [
        "MD&A / Financial Commentary",
        "Informational / Commercial",
        "P1",
        "/blog",
        "Live (related)",
        "Board/close workflow language; not a separate SKU.",
    ],
    [
        "Board Deck Automation",
        "Transactional",
        "P1",
        "/book-demo",
        "Live (CTA)",
        "Ads-leaning transactional modifier on board reporting.",
    ],
    [
        "Explainable AI for Finance",
        "Informational / Commercial",
        "P2",
        "/",
        "Live (implicit)",
        "Supporting trust angle for Copilot; not primary volume wager.",
    ],
    [
        "AI OS / AI Operating System for SaaS Finance",
        "Brand / Positioning",
        "Brand only",
        "/blog/ai-operating-system-for-saas-finance",
        "Live",
        "BRAND POSITIONING — NOT SEO NORTH STAR. Keep post; do not optimize site around this.",
    ],
    [
        "Finance OS",
        "Brand / Positioning",
        "Brand only",
        "/blog/finance-os-vs-fpa-software",
        "Live",
        "BRAND / CATEGORY NARRATIVE — NOT SEO NORTH STAR. Secondary to FP&A/board/ARR.",
    ],
    [
        "NRR / GRR Literacy",
        "Informational",
        "P0",
        "/blog/grr-vs-nrr",
        "Live or draft (see tracker)",
        "From SEO_Target_Keywords_2026-08 tracker — pillar + glossary support.",
    ],
    [
        "Billing vs CRM ARR",
        "Problem / Informational",
        "P1",
        "/blog/billing-vs-crm-arr",
        "Planned",
        "Content gap called out in Aug 2026 target tracker.",
    ],
]

KEYWORDS = [
    # term, theme, intent, channel, priority, negative?, notes
    ["SaaS FP&A software", "SaaS FP&A Software", "Transactional", "Both", "P0", "No", "Primary ads + SEO"],
    ["FP&A software for SaaS", "SaaS FP&A Software", "Transactional", "Both", "P0", "No", "Close variant"],
    ["SaaS financial intelligence", "SaaS Financial Intelligence", "Commercial", "Both", "P0", "No", "Category umbrella"],
    ["financial intelligence platform", "SaaS Financial Intelligence", "Commercial", "Ads", "P0", "No", "Qualify with SaaS in ad copy"],
    ["SaaS board reporting", "Board Reporting", "Commercial", "Both", "P0", "No", "Live blog target"],
    ["board reporting software", "Board Reporting", "Transactional", "Ads", "P0", "No", "Strong buy intent"],
    ["SaaS board pack", "Board Reporting", "Problem", "SEO", "P1", "No", "Watch GSC pack vs reporting split"],
    ["executive reporting software", "Executive Reporting", "Transactional", "Ads", "P1", "No", "Avoid generic BI dashboards"],
    ["ARR reporting", "ARR Reporting", "Commercial", "Both", "P0", "No", "One-liner theme"],
    ["ARR reporting software", "ARR Reporting", "Transactional", "Ads", "P0", "No", "Ads modifier"],
    ["ARR waterfall", "ARR Reporting", "Informational", "Both", "P0", "No", "Live cornerstone blog"],
    ["ARR waterfall vs GAAP revenue", "ARR Reporting", "Informational", "SEO", "P0", "No", "Live comparison post"],
    ["ARR governance", "ARR Reporting", "Informational", "SEO", "P0", "No", "Live methodology post"],
    ["billing vs CRM ARR", "Billing vs CRM ARR", "Problem", "SEO", "P1", "No", "Planned post — high pain"],
    ["NRR vs GRR", "NRR / GRR Literacy", "Informational", "SEO", "P0", "No", "Pillar + glossary"],
    ["net revenue retention SaaS", "NRR / GRR Literacy", "Informational", "SEO", "P1", "No", "Glossary + pillar"],
    ["gross revenue retention SaaS", "NRR / GRR Literacy", "Informational", "SEO", "P1", "No", "Glossary + pillar"],
    ["AI CFO copilot", "AI CFO Copilot", "Commercial", "Both", "P0", "No", "Ads-first; improve landing"],
    ["AI CFO", "AI CFO Copilot", "Commercial", "Ads", "P1", "No", "Broader — watch junk intent"],
    ["explainable AI for finance", "Explainable AI for Finance", "Commercial", "Both", "P2", "No", "Trust supporting"],
    ["SaaS cash forecast", "Cash Forecasting", "Informational", "SEO", "P1", "No", "Planned post"],
    ["SaaS forecasting software", "Cash Forecasting", "Transactional", "Ads", "P1", "No", "Cash + revenue forecast theme"],
    ["cash forecasting SaaS", "Cash Forecasting", "Commercial", "Both", "P1", "No", ""],
    ["scenario planning FP&A", "Scenario Analysis", "Commercial", "Both", "P1", "No", "One-liner theme"],
    ["scenario analysis SaaS", "Scenario Analysis", "Commercial", "Both", "P1", "No", ""],
    ["variance analysis software", "Variance Analysis", "Transactional", "Ads", "P2", "No", ""],
    ["MD&A reporting", "MD&A / Financial Commentary", "Informational", "SEO", "P1", "No", ""],
    ["board deck automation", "Board Deck Automation", "Transactional", "Ads", "P1", "No", ""],
    ["SaaS KPI reporting", "SaaS KPI Reporting", "Commercial", "Both", "P2", "No", ""],
    ["pipeline reporting SaaS", "Pipeline Reporting", "Commercial", "SEO", "P2", "No", ""],
    ["revenue forecasting software", "Revenue Forecasting", "Transactional", "Ads", "P1", "No", ""],
    ["AI operating system for SaaS finance", "AI OS / AI Operating System for SaaS Finance", "Brand", "SEO", "Brand only", "No", "BRAND POSITIONING — NOT SEO NORTH STAR"],
    ["finance operating system", "Finance OS", "Brand", "SEO", "Brand only", "No", "BRAND / NARRATIVE — NOT SEO NORTH STAR"],
    ["Finance OS vs FP&A software", "Finance OS", "Brand", "SEO", "Brand only", "No", "Thought-leadership only"],
    ["SMPL.ai", "SaaS Financial Intelligence", "Brand", "Ads", "P0", "No", "Brand exact / phrase"],
    ["SMPL AI finance", "SaaS Financial Intelligence", "Brand", "Ads", "P1", "No", "Brand variant"],
    # Explicit negative rows also listed on Negatives sheet
    ["bookkeeping", "—", "Wrong intent", "Ads", "—", "Yes", "Negative — not our product"],
    ["payroll", "—", "Wrong intent", "Ads", "—", "Yes", "Negative — not our product"],
    ["tax software", "—", "Wrong intent", "Ads", "—", "Yes", "Negative"],
    ["AP automation", "—", "Wrong intent", "Ads", "—", "Yes", "Negative"],
    ["ChatGPT for finance", "—", "Wrong intent", "Ads", "—", "Yes", "Negative — generic chatbot intent"],
    ["QuickBooks", "—", "Wrong intent", "Ads", "—", "Yes", "Negative as primary match; SoR not competitor replace"],
    ["accounting software", "—", "Wrong intent", "Ads", "—", "Yes", "Negative as primary — GL/ERP intent"],
]

NEGATIVES = [
    # term, match type suggestion, reason, apply to
    ["bookkeeping", "Broad / Phrase", "Wrong product — SMB bookkeeping", "All search campaigns"],
    ["bookkeeper", "Broad / Phrase", "Service-provider intent", "All search campaigns"],
    ["payroll", "Broad / Phrase", "Wrong product", "All search campaigns"],
    ["payroll software", "Phrase / Exact", "Wrong product", "All search campaigns"],
    ["tax", "Broad (monitor)", "Tax filing / tax software intent", "All search campaigns"],
    ["tax software", "Phrase / Exact", "Wrong product", "All search campaigns"],
    ["AP automation", "Phrase", "AP/invoice ops — not FP&A intelligence", "All search campaigns"],
    ["accounts payable", "Phrase", "AP ops intent", "All search campaigns"],
    ["invoice OCR", "Phrase", "Document capture intent", "All search campaigns"],
    ["QuickBooks", "Phrase / Exact", "SoR / bookkeeping; we do not replace", "All search campaigns"],
    ["Xero", "Phrase / Exact", "Accounting SoR intent", "All search campaigns"],
    ["accounting software", "Phrase", "GL replacement intent", "All search campaigns"],
    ["ChatGPT", "Broad / Phrase", "Generic chatbot wrappers", "All search campaigns"],
    ["ChatGPT for finance", "Phrase / Exact", "Wrong AI intent", "All search campaigns"],
    ["free", "Broad (monitor)", "Free-tool tire-kickers — use if CPA suffers", "All search campaigns"],
    ["template", "Broad (monitor)", "DIY template seekers — keep if blog SEO needs the traffic unpaid", "Ads only (optional)"],
    ["jobs", "Broad", "Hiring intent", "All search campaigns"],
    ["salary", "Broad", "Jobs/comp intent", "All search campaigns"],
    ["course", "Broad", "Education product intent", "All search campaigns"],
    ["certification", "Broad", "Training intent; also avoid SOC 2 certified claims in copy", "All search campaigns"],
]


def main() -> None:
    wb = Workbook()

    # --- Themes ---
    ws_t = wb.active
    ws_t.title = "Themes"
    write_table(
        ws_t,
        [
            "Theme name",
            "Intent type",
            "Priority",
            "Primary URL",
            "URL status (live / planned)",
            "Notes",
        ],
        THEMES,
    )
    highlight_priority(ws_t, 3, {"AI OS / AI Operating System for SaaS Finance", "Finance OS"})

    # --- Keywords ---
    ws_k = wb.create_sheet("Keywords")
    write_table(
        ws_k,
        [
            "Term",
            "Theme",
            "Intent",
            "Channel (SEO / Ads / Both)",
            "Priority (P0 / P1 / P2)",
            "Negative?",
            "Notes",
        ],
        KEYWORDS,
    )
    for row in ws_k.iter_rows(min_row=2, max_row=ws_k.max_row):
        pri = str(row[4].value or "")
        if pri == "P0":
            row[4].fill = P0_FILL
        if pri.lower().startswith("brand"):
            row[4].fill = BRAND_FILL
        if str(row[5].value or "").lower() == "yes":
            row[5].fill = NEG_FILL
            row[0].fill = NEG_FILL

    # --- Negatives ---
    ws_n = wb.create_sheet("Negatives")
    write_table(
        ws_n,
        ["Negative term", "Match type (suggestion)", "Reason", "Apply to"],
        NEGATIVES,
    )
    for row in ws_n.iter_rows(min_row=2, max_row=ws_n.max_row):
        row[0].fill = NEG_FILL

    # --- Review ---
    ws_r = wb.create_sheet("Review")
    ws_r["A1"] = "SEO / Ads Keyword Strategy — Review summary (Aug 2026)"
    ws_r["A1"].font = TITLE_FONT
    review_lines = [
        "",
        "Purpose: Shared intent-theme inventory for organic SEO + Google Ads. Qualitative priority only — no invented volumes.",
        "Related: SEO_Target_Keywords_2026-08.* (URL tracker) · SEO_Keywords_Before_After_2026-08.xlsx (history only).",
        "Not SOC 2 certified. Internal marketing use.",
        "",
        "=== PRIORITIZE — ORGANIC (SEO) ===",
        "1. ARR methodology cluster (waterfall, GAAP, governance, NRR/GRR; then billing vs CRM ARR).",
        "2. Board reporting / board pack (live post + homepage theme; internal links).",
        "3. SaaS FP&A / financial intelligence on conversion pages (/, /pricing, /book-demo).",
        "4. Cash forecast + scenario/variance content (planned bridges from board pack).",
        "5. Glossary entry points (NRR, GRR, waterfall, ARR) feeding pillars.",
        "",
        "=== PRIORITIZE — PAID (ADS) ===",
        "1. SaaS FP&A software / FP&A for SaaS.",
        "2. Board reporting software / SaaS board reporting.",
        "3. ARR reporting software (+ careful ARR waterfall).",
        "4. AI CFO Copilot / explainable AI for SaaS finance.",
        "5. Financial intelligence platform (SaaS-qualified ad copy).",
        "Conversion goal: Submit lead form (demo/quote). Page views secondary only.",
        "",
        "=== GAPS ===",
        "- Billing vs CRM ARR post (planned).",
        "- SaaS cash forecast post (planned).",
        "- Dedicated AI CFO Copilot landing for ads message match.",
        "- Scenario / variance landing still thin.",
        "- Validate with GSC + Ads search terms (no fabricated volumes here).",
        "",
        "=== AVOID / NEGATIVES ===",
        "- Bookkeeping, payroll, tax, AP automation, invoice OCR, QuickBooks-as-primary, accounting software-as-primary.",
        "- Generic ChatGPT / chatbot-for-finance intent.",
        "- Positioning as NetSuite/Intacct/Salesforce/Stripe replacement (we connect / load data).",
        "- SOC 2 certified claims (not certified).",
        "- AI OS / Finance OS as SEO or ads north star (brand/positioning only).",
        "",
        "=== BRAND SLOGANS (NOT NORTH STAR) ===",
        "- AI operating system for SaaS finance — brand/positioning; keep thought-leadership post.",
        "- Finance OS — category narrative; secondary to concrete FP&A / board / ARR queries.",
        "",
        "=== PRODUCT ONE-LINERS (KEEP) ===",
        "SaaS Financial Intelligence · SaaS FP&A Software · ARR Reporting · Board Reporting ·",
        "Cash Forecasting · Scenario Analysis · AI CFO Copilot · Executive Reporting ·",
        "Pipeline Reporting · Revenue Forecasting · SaaS KPI Reporting · Variance Analysis ·",
        "MD&A / Financial Commentary · Board Deck Automation · Explainable AI for Finance",
        "",
        "=== MATT REVIEW CHECKLIST ===",
        "[ ] Confirm P0 themes for ads launch",
        "[ ] Load Negatives sheet into Google Ads before scaling",
        "[ ] Keep AI OS / Finance OS off primary keyword lists",
        "[ ] Approve next content briefs: billing vs CRM ARR; SaaS cash forecast",
        "[ ] Decide whether to add thin AI CFO Copilot landing for ads",
        "",
        f"Output: {OUT.as_posix()}",
    ]
    for i, line in enumerate(review_lines, start=2):
        ws_r[f"A{i}"] = line
        ws_r[f"A{i}"].alignment = Alignment(wrap_text=True)
    ws_r.column_dimensions["A"].width = 110

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Themes: {len(THEMES)}; Keywords: {len(KEYWORDS)}; Negatives: {len(NEGATIVES)}")


if __name__ == "__main__":
    main()
