"""Build founder brainstorm Excel for Google Sheets upload.

Reads docs/marketing/SEO_Keyword_Strategy_Intent_Themes.xlsx and writes
docs/marketing/SEO_Keyword_Founder_Brainstorm.xlsx with collaboration columns.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "docs" / "marketing" / "SEO_Keyword_Strategy_Intent_Themes.xlsx"
OUT = ROOT / "docs" / "marketing" / "SEO_Keyword_Founder_Brainstorm.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
BODY_FONT = Font(size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALT_FILL = PatternFill("solid", fgColor="F5F8FC")
HINT_FILL = PatternFill("solid", fgColor="FFF8E7")


def style_header(ws, row: int = 1, cols: int = 1) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN


def autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def freeze_and_filter(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def add_blank_rows(ws, start_row: int, n: int, cols: int) -> None:
    for r in range(start_row, start_row + n):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c, "")
            cell.border = THIN
            cell.alignment = WRAP
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_FILL


def read_src_rows(sheet_name: str) -> list[tuple]:
    if not SRC.exists():
        raise FileNotFoundError(f"Missing source strategy workbook: {SRC}")
    wb = load_workbook(SRC, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not in {SRC.name}")
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        if not any(v is not None and str(v).strip() for v in row):
            continue
        rows.append(tuple("" if v is None else str(v).strip() for v in row))
    return rows


def build_how_to(wb: Workbook) -> None:
    ws = wb.create_sheet("How to use", 0)
    lines = [
        ("SEO / Ads keyword brainstorm — founder worksheet", True),
        ("", False),
        ("Purpose", True),
        (
            "Shared working list for founders to review themes/keywords, vote, add ideas, "
            "and decide Keep / Cut / Needs research. Prefills from the Aug 2026 intent-theme strategy.",
            False,
        ),
        ("", False),
        ("Recommended workflow", True),
        (
            "1) Optional: dump messy ideas in FigJam or Miro first (sticky notes, no structure).",
            False,
        ),
        (
            "2) Park winners into THIS Google Sheet (Themes / Keywords / Parking lot).",
            False,
        ),
        (
            "3) Use comments + Owner + Status columns to drive the discussion.",
            False,
        ),
        (
            "4) After founders agree, update the git strategy brief "
            "(SEO_Keyword_Strategy_Intent_Themes.md) as the approved snapshot — "
            "do not use git markdown as the live brainstorm surface.",
            False,
        ),
        ("", False),
        ("Upload to Google Sheets (Matt / whoever owns Drive)", True),
        ("1. Open Google Drive → New → File upload → select this .xlsx file.", False),
        ("2. Right-click the uploaded file → Open with → Google Sheets.", False),
        (
            "3. File → Share → add founders (Commenter or Editor). Turn on notifications if useful.",
            False,
        ),
        (
            "4. Optionally File → Make a copy for a dated working session; keep one master sheet.",
            False,
        ),
        (
            "5. After decisions: copy Keep rows back into the strategy snapshot (or ask eng to sync).",
            False,
        ),
        ("", False),
        ("Tabs", True),
        ("How to use — these instructions.", False),
        (
            "Themes — product/theme one-liners. Vote 1–3 (1 = must prioritize). "
            "Status: Idea / Keep / Cut / Needs research. Fill Founder notes + Owner.",
            False,
        ),
        (
            "Keywords — concrete terms tied to a theme. Mark Founder idea? = Y when someone adds a new term. "
            "Status same as Themes.",
            False,
        ),
        (
            "Parking lot — blank capture for new themes/keywords during the meeting. "
            "Move winners into Themes or Keywords and set Status.",
            False,
        ),
        (
            "Avoid Negatives — wrong-intent terms (especially ads). Add more as you find junk queries.",
            False,
        ),
        (
            "Decisions — log what you agreed (date, decision, who, follow-up). This becomes the sync checklist for the strategy snapshot.",
            False,
        ),
        ("", False),
        ("Column cheat sheet", True),
        ("Vote — 1 (high) / 2 (medium) / 3 (low). Leave blank until discussed.", False),
        ("Owner — initials of who will research or champion the theme/term.", False),
        (
            "Status — Idea (new) / Keep (in strategy) / Cut (drop) / Needs research (validate in GSC/Ads).",
            False,
        ),
        ("Channel — SEO / Ads / Both / Brand only.", False),
        ("Priority (seed) — P0/P1/P2 from current strategy; change Vote if you disagree.", False),
        ("", False),
        ("North-star reminder", True),
        (
            "Rank and advertise for buyer queries (SaaS FP&A, board reporting, ARR, financial intelligence). "
            "AI OS / Finance OS = brand positioning only — not the SEO/Ads north star.",
            False,
        ),
        ("", False),
        (
            "Rebuild: python scripts/_build_seo_founder_brainstorm_xlsx.py "
            "(overwrites this file from SEO_Keyword_Strategy_Intent_Themes.xlsx).",
            False,
        ),
        ("Internal marketing use only. Not SOC 2 certified. No fabricated search volumes.", False),
    ]
    ws.column_dimensions["A"].width = 110
    for i, (text, bold) in enumerate(lines, 1):
        cell = ws.cell(i, 1, text)
        cell.font = SECTION_FONT if bold else BODY_FONT
        cell.alignment = WRAP
        if bold and text:
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
    ws.row_dimensions[1].height = 22


def build_themes(wb: Workbook) -> None:
    ws = wb.create_sheet("Themes")
    headers = [
        "Theme",
        "Intent type",
        "Priority (seed)",
        "Primary URL",
        "URL status",
        "Strategy notes",
        "Founder notes",
        "Vote (1-3)",
        "Owner",
        "Status",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for row in read_src_rows("Themes"):
        # Theme name | Intent type | Priority | Primary URL | URL status | Notes
        theme, intent, priority, url, url_status, notes = (list(row) + [""] * 6)[:6]
        status = "Keep"
        if "brand" in (priority or "").lower():
            status = "Keep"  # brand rows stay visible; founders can Cut from SEO wager
        ws.append(
            [
                theme,
                intent,
                priority,
                url,
                url_status,
                notes,
                "",
                "",
                "",
                status,
            ]
        )

    # Extra blank idea rows
    start = ws.max_row + 1
    add_blank_rows(ws, start, 12, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.alignment = WRAP
            cell.border = THIN
            if c in (7, 8, 9, 10) and r >= start:
                cell.fill = HINT_FILL

    dv = DataValidation(
        type="list",
        formula1='"Idea,Keep,Cut,Needs research"',
        allow_blank=True,
    )
    dv.error = "Pick Idea, Keep, Cut, or Needs research"
    dv.errorTitle = "Status"
    ws.add_data_validation(dv)
    dv.add(f"J2:J{ws.max_row}")

    dv_vote = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    ws.add_data_validation(dv_vote)
    dv_vote.add(f"H2:H{ws.max_row}")

    autosize(
        ws,
        {
            1: 38,
            2: 28,
            3: 14,
            4: 36,
            5: 22,
            6: 48,
            7: 32,
            8: 12,
            9: 12,
            10: 16,
        },
    )
    freeze_and_filter(ws)
    ws.row_dimensions[1].height = 30


def build_keywords(wb: Workbook) -> None:
    ws = wb.create_sheet("Keywords")
    headers = [
        "Term",
        "Theme",
        "Intent",
        "Channel",
        "Priority (seed)",
        "Founder idea? (Y/N)",
        "Founder notes",
        "Vote (1-3)",
        "Owner",
        "Status",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for row in read_src_rows("Keywords"):
        # Term | Theme | Intent | Channel | Priority | Negative? | Notes
        term, theme, intent, channel, priority, negative, notes = (list(row) + [""] * 7)[:7]
        if (negative or "").lower() == "yes":
            continue  # negatives live on Avoid sheet
        ws.append(
            [
                term,
                theme,
                intent,
                channel,
                priority,
                "N",
                notes,
                "",
                "",
                "Keep",
            ]
        )

    start = ws.max_row + 1
    add_blank_rows(ws, start, 20, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.alignment = WRAP
            cell.border = THIN
            if r >= start and c in (1, 6, 7, 8, 9, 10):
                cell.fill = HINT_FILL

    dv_status = DataValidation(
        type="list",
        formula1='"Idea,Keep,Cut,Needs research"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"J2:J{ws.max_row}")

    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"F2:F{ws.max_row}")

    dv_vote = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    ws.add_data_validation(dv_vote)
    dv_vote.add(f"H2:H{ws.max_row}")

    autosize(
        ws,
        {
            1: 36,
            2: 34,
            3: 18,
            4: 12,
            5: 14,
            6: 16,
            7: 36,
            8: 12,
            9: 12,
            10: 16,
        },
    )
    freeze_and_filter(ws)
    ws.row_dimensions[1].height = 30


def build_parking_lot(wb: Workbook) -> None:
    ws = wb.create_sheet("Parking lot")
    headers = [
        "Theme or keyword?",
        "Idea (theme / keyword text)",
        "Intent guess",
        "Channel guess",
        "Who suggested",
        "Why it matters",
        "Move to Themes/Keywords?",
        "Status",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    # Seed a few prompt rows from known gaps so the sheet isn't empty
    seeds = [
        (
            "Theme",
            "Billing vs CRM ARR",
            "Problem / Informational",
            "SEO",
            "(seed)",
            "High reconciliation pain; planned pillar in strategy",
            "",
            "Needs research",
        ),
        (
            "Theme",
            "SaaS cash forecast",
            "Informational / Commercial",
            "Both",
            "(seed)",
            "Bridge from board ARR–cash–P&L post; soft demo CTA",
            "",
            "Needs research",
        ),
        (
            "Keyword",
            "(your idea here)",
            "",
            "",
            "",
            "",
            "",
            "Idea",
        ),
    ]
    for s in seeds:
        ws.append(list(s))
    start = ws.max_row + 1
    add_blank_rows(ws, start, 25, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.alignment = WRAP
            cell.border = THIN
            if r >= 4:  # blank capture area
                cell.fill = HINT_FILL

    dv = DataValidation(
        type="list",
        formula1='"Idea,Keep,Cut,Needs research"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"H2:H{ws.max_row}")

    dv_type = DataValidation(type="list", formula1='"Theme,Keyword"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add(f"A2:A{ws.max_row}")

    autosize(ws, {1: 18, 2: 36, 3: 24, 4: 14, 5: 16, 6: 42, 7: 22, 8: 16})
    freeze_and_filter(ws)
    ws.row_dimensions[1].height = 30


def build_avoid(wb: Workbook) -> None:
    ws = wb.create_sheet("Avoid Negatives")
    headers = [
        "Negative term",
        "Match type (suggestion)",
        "Reason",
        "Apply to",
        "Founder notes",
        "Keep as negative? (Y/N)",
        "Who added",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for row in read_src_rows("Negatives"):
        term, match, reason, apply = (list(row) + [""] * 4)[:4]
        ws.append([term, match, reason, apply, "", "Y", ""])

    # Extra guidance rows from Review avoid list (SoR replacement / SOC2 / brand)
    extras = [
        (
            "NetSuite replacement",
            "Phrase",
            "We connect; we do not replace SoR",
            "All search campaigns",
            "",
            "Y",
            "(seed)",
        ),
        (
            "Sage Intacct replacement",
            "Phrase",
            "We connect; we do not replace SoR",
            "All search campaigns",
            "",
            "Y",
            "(seed)",
        ),
        (
            "Salesforce replacement",
            "Phrase",
            "We connect; we do not replace SoR",
            "All search campaigns",
            "",
            "Y",
            "(seed)",
        ),
        (
            "SOC 2 certified",
            "Phrase / Exact",
            "Not certified — do not claim in ads/SEO copy",
            "All campaigns + copy review",
            "",
            "Y",
            "(seed)",
        ),
    ]
    for e in extras:
        ws.append(list(e))

    start = ws.max_row + 1
    add_blank_rows(ws, start, 15, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.alignment = WRAP
            cell.border = THIN
            if r >= start:
                cell.fill = HINT_FILL

    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"F2:F{ws.max_row}")

    autosize(ws, {1: 28, 2: 20, 3: 48, 4: 28, 5: 28, 6: 18, 7: 12})
    freeze_and_filter(ws)
    ws.row_dimensions[1].height = 30


def build_decisions(wb: Workbook) -> None:
    ws = wb.create_sheet("Decisions")
    headers = [
        "Date",
        "Decision",
        "Who",
        "Related theme / keyword",
        "Follow-up",
        "Done? (Y/N)",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    # Starter checklist from strategy Review
    starters = [
        (
            "",
            "Confirm P0 themes for ads launch",
            "",
            "SaaS FP&A; Board; ARR; Financial Intelligence; AI CFO Copilot",
            "Align Ads campaigns to Keep rows",
            "N",
        ),
        (
            "",
            "Load Negatives into Google Ads before scaling spend",
            "",
            "Avoid Negatives",
            "Import negatives sheet",
            "N",
        ),
        (
            "",
            "Keep AI OS / Finance OS off primary keyword lists",
            "",
            "AI OS; Finance OS",
            "Brand-only treatment",
            "N",
        ),
        (
            "",
            "Approve next content briefs: billing vs CRM ARR; SaaS cash forecast",
            "",
            "Parking lot seeds",
            "Assign Owner on Themes",
            "N",
        ),
        (
            "",
            "Decide whether to add thin AI CFO Copilot landing for ads",
            "",
            "AI CFO Copilot",
            "Message match / Quality Score",
            "N",
        ),
    ]
    for s in starters:
        ws.append(list(s))
    start = ws.max_row + 1
    add_blank_rows(ws, start, 20, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.alignment = WRAP
            cell.border = THIN
            if r >= start:
                cell.fill = HINT_FILL

    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"F2:F{ws.max_row}")

    autosize(ws, {1: 12, 2: 56, 3: 16, 4: 40, 5: 32, 6: 12})
    freeze_and_filter(ws)
    ws.row_dimensions[1].height = 30


def main() -> None:
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    build_how_to(wb)
    build_themes(wb)
    build_keywords(wb)
    build_parking_lot(wb)
    build_avoid(wb)
    build_decisions(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
